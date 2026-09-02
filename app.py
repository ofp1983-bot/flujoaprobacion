import streamlit as st
import sqlite3
import uuid
import os
import io
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, timezone, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from pypdf import PdfWriter, PdfReader

# ==========================================
# CONFIGURACIÓN GENERAL Y CREDENCIALES
# ==========================================
DB_FILE = "flujo_documental.db"

# Zona horaria de Colombia (UTC-5)
ZONA_COL = timezone(timedelta(hours=-5))

# CONFIGURACIÓN NEXTCLOUD
NC_USER = "gdocumental@insdeportescajica.gov.co"
NC_PASS = "7Lb2n-WdwWQ-SzM5n-JYA5J-Px5Ro" 
# Asegúrate de usar la URL correcta para la API WebDAV
NC_URL = f"https://cloud.insdeportescajica.gov.co/remote.php/dav/files/19C87196-1654-4F6B-A835-7255DBC00FF1/DIRECCION/APROBACIONES"
AUTH = HTTPBasicAuth(NC_USER, NC_PASS)

# URL base para los enlaces
BASE_URL = "http://localhost:8501"

# ==========================================
# FUNCIONES DE BASE DE DATOS
# ==========================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS documentos (
                    id TEXT PRIMARY KEY, 
                    nombre_archivo TEXT, 
                    ruta_archivo TEXT, 
                    observacion_inicial TEXT, 
                    estado TEXT, 
                    total_revisores INTEGER, 
                    total_aprobadores INTEGER, 
                    fecha_creacion TEXT,
                    nombre_elaborador TEXT, 
                    correo_elaborador TEXT, 
                    cargo_elaborador TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS firmas_flujo (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, 
                    doc_id TEXT, 
                    token TEXT UNIQUE,
                    rol TEXT, 
                    correo_esperado TEXT,
                    nombre TEXT, 
                    cargo TEXT, 
                    correo TEXT, 
                    decision TEXT, 
                    observacion TEXT, 
                    fecha_hora TEXT)''')
    conn.commit()
    conn.close()

init_db()

# ==========================================
# FUNCIONES DE NEXTCLOUD (WEBDAV)
# ==========================================
def subir_a_nextcloud(archivo_bytes, nombre_archivo):
    url = f"{NC_URL}/{nombre_archivo}"
    try:
        respuesta = requests.put(url, data=archivo_bytes, auth=AUTH)
        if respuesta.status_code in [201, 204]:
            return True
        else:
            st.error(f"Fallo en Nextcloud - Código HTTP {respuesta.status_code}: {respuesta.text}")
            return False
    except requests.exceptions.RequestException as e:
        st.error(f"Error de conexión con el servidor: {e}")
        return False

def descargar_de_nextcloud(nombre_archivo):
    url = f"{NC_URL}/{nombre_archivo}"
    respuesta = requests.get(url, auth=AUTH)
    if respuesta.status_code == 200:
        return respuesta.content
    return None

# ==========================================
# FUNCIONES DE GENERACIÓN PDF Y CIERRE
# ==========================================
def generar_hoja_control_temporal(doc_id):
    with sqlite3.connect(DB_FILE) as conn:
        c = conn.cursor()
        c.execute("""SELECT rol, nombre, cargo, correo, decision, fecha_hora 
                     FROM firmas_flujo 
                     WHERE doc_id = ? AND decision = 'APROBADO' 
                     ORDER BY id ASC""", (doc_id,))
        firmas_definitivas = c.fetchall()
        
        c.execute("""SELECT nombre_elaborador, cargo_elaborador, correo_elaborador, fecha_creacion 
                     FROM documentos WHERE id = ?""", (doc_id,))
        doc_info = c.fetchone()

    temp_hoja = f"hoja_control_{doc_id}.pdf"
    doc = SimpleDocTemplate(temp_hoja, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    estilo_celda = ParagraphStyle(
        'EstiloCelda',
        parent=styles['Normal'],
        fontSize=7,
        alignment=TA_CENTER,
        leading=8
    )

    story = []
    story.append(Paragraph("<b>MATRIZ DE CONTROL Y APROBACIÓN DOCUMENTAL</b>", styles['Title']))
    story.append(Spacer(1, 20))

    data = [["Rol", "Nombre Completo", "Cargo", "Correo", "Estado", "Fecha y Hora"]]
    
    def ajustar_texto(texto):
        return Paragraph(str(texto), estilo_celda)

    data.append([
        ajustar_texto("Elaborador"), 
        ajustar_texto(doc_info[0]), 
        ajustar_texto(doc_info[1]), 
        ajustar_texto(doc_info[2]), 
        ajustar_texto("ELABORADO"), 
        ajustar_texto(doc_info[3])
    ])

    for f in firmas_definitivas:
        data.append([
            ajustar_texto(f[0]), ajustar_texto(f[1]), ajustar_texto(f[2]), 
            ajustar_texto(f[3]), ajustar_texto(f[4]), ajustar_texto(f[5])
        ])

    t = Table(data, colWidths=[60, 110, 110, 120, 65, 87])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1A365D")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    
    story.append(t)
    doc.build(story)
    return temp_hoja

def procesar_cierre_documento(doc_id_input, doc_nombre_nextcloud):
    ruta_hoja_temporal = generar_hoja_control_temporal(doc_id_input)
    pdf_original_bytes = descargar_de_nextcloud(doc_nombre_nextcloud)
    
    if not pdf_original_bytes:
        return None, None

    reader_orig = PdfReader(io.BytesIO(pdf_original_bytes))
    reader_hoja = PdfReader(ruta_hoja_temporal)
    writer = PdfWriter()

    for page in reader_orig.pages:
        writer.add_page(page)
    for page in reader_hoja.pages:
        writer.add_page(page)

    pdf_final_io = io.BytesIO()
    writer.write(pdf_final_io)
    pdf_final_bytes = pdf_final_io.getvalue()

    nombre_final_nc = doc_nombre_nextcloud.replace(".pdf", "_FINAL.pdf")
    subir_a_nextcloud(pdf_final_bytes, nombre_final_nc)
    
    if os.path.exists(ruta_hoja_temporal):
        os.remove(ruta_hoja_temporal)
        
    return pdf_final_bytes, nombre_final_nc

# ==========================================
# INTERFAZ DE USUARIO - STREAMLIT
# ==========================================
st.set_page_config(page_title="Gestión y Aprobación Documental", layout="wide")

params = st.query_params
token = params.get("token", None)

if token:
    # ----------------- VISTA REVISOR / APROBADOR -----------------
    with sqlite3.connect(DB_FILE) as conn:
        c = conn.cursor()
        c.execute("SELECT doc_id, rol, decision, correo_esperado FROM firmas_flujo WHERE token = ?", (token,))
        registro = c.fetchone()

        if not registro:
            st.error("Enlace no válido o expirado.")
        elif registro[2] != "PENDIENTE":
            st.info(f"Ya has completado esta gestión. Estado actual: {registro[2]}")
        else:
            doc_id, rol, _, correo_esperado = registro
            c.execute("SELECT nombre_archivo, ruta_archivo, observacion_inicial FROM documentos WHERE id = ?", (doc_id,))
            doc = c.fetchone()
            
            st.title(f"Gestión de {rol} de Documento")
            st.write(f"**Archivo:** {doc[0]}")
            nota_formateada = doc[2].replace('\n', '  \n')
            st.write(f"**Notas del flujo:**  \n{nota_formateada}")

            nombre_nextcloud = doc[1]
            pdf_bytes = descargar_de_nextcloud(nombre_nextcloud)

            if pdf_bytes:
                st.download_button("📥 Descargar documento para revisión", data=pdf_bytes, file_name=doc[0], mime="application/pdf", key="btn_dl_revision")
            else:
                st.error("Error: El documento no se encontró en el repositorio de Nextcloud.")

            st.divider()
            decision = st.radio("Indica tu decisión:", ["Aprobar", "Rechazar/Devolver"])

            with st.form("form_firma"):
                nombre = st.text_input("Nombre completo:")
                cargo = st.text_input("Cargo:")
                correo = st.text_input("Correo electrónico institucional:")
                observaciones = st.text_area("Observaciones (obligatorio si rechaza/devuelve):", key="obs_firma")
                submit = st.form_submit_button("Confirmar y Registrar")

                if submit:
                    if correo.strip().lower() != correo_esperado.strip().lower():
                        st.error("El correo electrónico no coincide con el asignado para este enlace.")
                    elif not nombre or not cargo or not correo:
                        st.error("Nombre, cargo y correo son obligatorios.")
                    elif decision == "Rechazar/Devolver" and not observaciones.strip():
                        st.error("Debe ingresar una observación para justificar la devolución.")
                    else:
                        now = datetime.now(ZONA_COL).strftime("%Y-%m-%d %H:%M:%S")
                        estado_decision = "APROBADO" if decision == "Aprobar" else "RECHAZADO"
                        
                        c.execute("""UPDATE firmas_flujo 
                                     SET nombre=?, cargo=?, correo=?, decision=?, observacion=?, fecha_hora=? 
                                     WHERE token=?""", (nombre, cargo, correo, estado_decision, observaciones, now, token))
                        
                        if estado_decision == "RECHAZADO":
                            nuevo_estado = "RECHAZADO_REV" if rol == "REVISOR" else "RECHAZADO_APR"
                            c.execute("UPDATE documentos SET estado = ? WHERE id = ?", (nuevo_estado, doc_id))
                        
                        conn.commit()
                        st.success("Acción registrada con éxito. Ya puedes cerrar esta ventana.")
else:
    # ----------------- PANEL PRINCIPAL -----------------
    st.sidebar.title("Menú de Gestión")
    menu = st.sidebar.radio("Opciones", ["1. Iniciar Nuevo Flujo", "2. Estado y Aprobadores", "3. Documentos Finalizados", "4. Respaldo de Base de Datos"])

    if menu == "1. Iniciar Nuevo Flujo":
        st.header("Carga de Documento e Inicio de Flujo")
        
        st.subheader("Datos del Elaborador")
        col1, col2, col3 = st.columns(3)
        with col1:
            nombre_elab = st.text_input("Tu Nombre Completo:")
        with col2:
            cargo_elab = st.text_input("Tu Cargo:")
        with col3:
            correo_elab = st.text_input("Tu Correo Institucional:")

        st.subheader("Configuración del Documento")
        archivo = st.file_uploader("Selecciona el documento (PDF)", type=["pdf"], key="archivo_inicio")
        num_revisores = st.number_input("Número de Revisores requeridos:", min_value=1, max_value=10, value=1, key="num_rev_inicio")
        
        st.write("**Correos institucionales autorizados para revisión:**")
        correos_revisores = []
        for i in range(num_revisores):
            correos_revisores.append(st.text_input(f"Correo del Revisor {i+1}:", key=f"rev_{i}"))
            
        observacion = st.text_area("Nota / Observaciones (Ej: Primera versión / Ajustes realizados):", key="obs_inicio")

        if st.button("Crear Flujo y Generar Enlaces", key="btn_crear_flujo"):
            if not all(correos_revisores):
                st.warning("Debes ingresar el correo de todos los revisores asignados.")
            elif archivo and observacion and nombre_elab and cargo_elab and correo_elab:
                doc_id = str(uuid.uuid4())[:8]
                nombre_guardado = f"{doc_id}_{archivo.name}"
                
                with st.spinner("Subiendo a Nextcloud..."):
                    exito_subida = subir_a_nextcloud(archivo.getvalue(), nombre_guardado)
                
                if exito_subida:
                    with sqlite3.connect(DB_FILE) as conn:
                        c = conn.cursor()
                        fecha_creacion = datetime.now(ZONA_COL).strftime("%Y-%m-%d %H:%M:%S")
                        
                        c.execute("""INSERT INTO documentos VALUES (?, ?, ?, ?, 'REVISION', ?, 0, ?, ?, ?, ?)""",
                                  (doc_id, archivo.name, nombre_guardado, observacion, num_revisores, 
                                   fecha_creacion, nombre_elab, correo_elab, cargo_elab))
                        
                        tokens = []
                        for correo_esp in correos_revisores:
                            tok = str(uuid.uuid4())
                            c.execute("""INSERT INTO firmas_flujo (doc_id, token, rol, correo_esperado, decision) 
                                         VALUES (?, ?, 'REVISOR', ?, 'PENDIENTE')""", (doc_id, tok, correo_esp.strip().lower()))
                            tokens.append(tok)

                        conn.commit()

                    st.success(f"Documento guardado en Nextcloud. Flujo creado (ID: {doc_id}).")
                    st.info("Comparte los siguientes enlaces con los revisores:")
                    for i, tok in enumerate(tokens, 1):
                        st.code(f"{BASE_URL}/?token={tok}", language="text")
            else:
                st.warning("Por favor, completa todos los campos obligatorios y adjunta un documento.")

    elif menu == "2. Estado y Aprobadores":
        st.header("Control de Estados y Asignación de Aprobadores")
        doc_id_input = st.text_input("Ingresa el ID del Documento (Ej: abc123df):")
        
        if doc_id_input:
            with sqlite3.connect(DB_FILE) as conn:
                c = conn.cursor()
                c.execute("SELECT nombre_archivo, ruta_archivo, estado, total_revisores, observacion_inicial FROM documentos WHERE id = ?", (doc_id_input,))
                doc = c.fetchone()

                if doc:
                    st.write(f"**Archivo original:** {doc[0]} | **Estado Actual:** `{doc[2]}`")
                    
                    c.execute("SELECT rol, nombre, cargo, decision, observacion, fecha_hora FROM firmas_flujo WHERE doc_id = ?", (doc_id_input,))
                    firmas = c.fetchall()
                    
                    if firmas:
                        df_firmas = pd.DataFrame(firmas, columns=["Rol", "Nombre", "Cargo", "Decisión", "Observación", "Fecha y Hora"])
                        df_firmas["Observación"] = df_firmas["Observación"].fillna("Sin observaciones")
                        st.dataframe(df_firmas, use_container_width=True, hide_index=True)

                        st.write("---")
                        st.subheader("Control y Seguimiento")
                        
                        c.execute("SELECT nombre_elaborador, cargo_elaborador, fecha_creacion FROM documentos WHERE id = ?", (doc_id_input,))
                        elab_data = c.fetchone()
                        
                        historial = [{
                            'Rol': 'Elaborador', 'Nombre': elab_data[0], 'Cargo': elab_data[1],
                            'Decisión': 'ELABORADO', 'Observación': 'Inicio del flujo / Envío a revisión', 'Fecha y Hora': elab_data[2]
                        }]
                        for f in firmas:
                            historial.append({'Rol': f[0], 'Nombre': f[1], 'Cargo': f[2], 'Decisión': f[3], 'Observación': f[4] if f[4] else "", 'Fecha y Hora': f[5]})
                            
                        wb = Workbook()
                        ws = wb.active
                        ws.title = f"Trazabilidad_{doc_id_input}"
                        ws.merge_cells('A1:F1')
                        title_cell = ws['A1']
                        title_cell.value = f"BITÁCORA DE TRAZABILIDAD DOCUMENTAL - ID: {doc_id_input}"
                        title_cell.font = Font(bold=True, color="FFFFFF", size=13)
                        title_cell.fill = PatternFill(start_color="1A365D", fill_type="solid")
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws.append([])
                        
                        for r in dataframe_to_rows(pd.DataFrame(historial), index=False, header=True):
                            ws.append(r)
                            
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
                        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=6)):
                            for row_idx, cell in enumerate(col_cells):
                                cell.border = thin_border
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                                if row_idx == 0:
                                    cell.font = Font(bold=True)
                                    cell.fill = header_fill
                                    
                        ws.column_dimensions['A'].width = 12
                        ws.column_dimensions['B'].width = 25
                        ws.column_dimensions['C'].width = 30
                        ws.column_dimensions['D'].width = 12
                        ws.column_dimensions['E'].width = 40
                        ws.column_dimensions['F'].width = 18

                        excel_io = io.BytesIO()
                        wb.save(excel_io)
                        excel_io.seek(0)
                        
                        st.download_button(
                            label="📥 Descargar Bitácora de Trazabilidad (.xlsx)",
                            data=excel_io,
                            file_name=f"Bitacora_Trazabilidad_{doc_id_input}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_dl_excel"
                        )

                    st.write("---")
                    c.execute("SELECT COUNT(*) FROM firmas_flujo WHERE doc_id = ? AND rol = 'REVISOR' AND decision = 'APROBADO'", (doc_id_input,))
                    rev_aprobados = c.fetchone()[0]

                    if doc[2] == 'REVISION' and rev_aprobados == doc[3]:
                        st.success("✅ Todos los revisores han aprobado el documento.")
                        st.subheader("Paso 2: Configurar Aprobadores")
                        num_apr = st.number_input("Número de Aprobadores requeridos:", min_value=1, max_value=10, value=1, key="num_apr_conf")
                        
                        st.write("**Correos institucionales autorizados para aprobación:**")
                        correos_aprobadores = []
                        for i in range(num_apr):
                            correos_aprobadores.append(st.text_input(f"Correo del Aprobador {i+1}:", key=f"apr_{i}"))
                        
                        if st.button("Generar Enlaces para Aprobadores", key="btn_gen_aprobadores"):
                            if not all(correos_aprobadores):
                                st.warning("Debes ingresar el correo de todos los aprobadores asignados.")
                            else:
                                tokens_apr = []
                                for correo_esp in correos_aprobadores:
                                    tok = str(uuid.uuid4())
                                    c.execute("INSERT INTO firmas_flujo (doc_id, token, rol, correo_esperado, decision) VALUES (?, ?, 'APROBADOR', ?, 'PENDIENTE')", (doc_id_input, tok, correo_esp.strip().lower()))
                                    tokens_apr.append(tok)
                                
                                c.execute("UPDATE documentos SET estado = 'APROBACION', total_aprobadores = ? WHERE id = ?", (num_apr, doc_id_input))
                                conn.commit()
                                st.success("Enlaces de aprobación generados con éxito:")
                                for tok in tokens_apr:
                                    st.code(f"{BASE_URL}/?token={tok}", language="text")
                                    
                    elif doc[2] == 'APROBACION':
                        c.execute("SELECT COUNT(*) FROM firmas_flujo WHERE doc_id = ? AND rol = 'APROBADOR' AND decision = 'APROBADO'", (doc_id_input,))
                        apr_aprobados = c.fetchone()[0]
                        c.execute("SELECT total_aprobadores FROM documentos WHERE id = ?", (doc_id_input,))
                        tot_apr = c.fetchone()[0]

                        if apr_aprobados == tot_apr:
                            st.success("🎉 ¡Todos los aprobadores han firmado! El flujo está listo para cerrarse.")
                            if st.button("Generar Documento Final con Hoja de Firmas", key="btn_gen_pdf_final"):
                                with st.spinner("Descargando, uniendo PDFs y subiendo a Nextcloud..."):
                                    pdf_final_bytes, nombre_final_nc = procesar_cierre_documento(doc_id_input, doc[1])
                                    if pdf_final_bytes:
                                        c.execute("UPDATE documentos SET estado = 'COMPLETADO' WHERE id = ?", (doc_id_input,))
                                        conn.commit()
                                        st.success(f"Documento final guardado en Nextcloud como: `{nombre_final_nc}`")
                                        st.download_button("📥 Descargar PDF Final", data=pdf_final_bytes, file_name=nombre_final_nc, mime="application/pdf", key="btn_dl_pdf_final")
                                    else:
                                        st.error("Error obteniendo el archivo original desde Nextcloud.")

                    elif doc[2] in ['RECHAZADO_REV', 'RECHAZADO_APR']:
                        st.error("❌ Este documento fue rechazado. El flujo se encuentra detenido.")
                        st.subheader("🔄 Reiniciar Flujo (Subir Correcciones)")
                        st.info(f"Al reiniciar, se conservará el ID **{doc_id_input}** para mantener toda la trazabilidad histórica.")
                        
                        nuevo_archivo = st.file_uploader("Sube el documento corregido (PDF)", type=["pdf"], key="archivo_reinicio")
                        nueva_obs = st.text_area("Nota sobre los ajustes realizados:", key="obs_reinicio")
                        nuevos_rev = st.number_input("Número de revisores para esta nueva ronda:", min_value=1, max_value=10, value=doc[3], key="num_rev_reinicio")
                        
                        st.write("**Correos institucionales autorizados para la nueva revisión:**")
                        correos_reinicio = []
                        for i in range(nuevos_rev):
                            correos_reinicio.append(st.text_input(f"Correo del Revisor {i+1}:", key=f"rein_{i}"))
                        
                        if st.button("Reiniciar Flujo y Generar Nuevos Enlaces", key="btn_reinicio_flujo"):
                            if not all(correos_reinicio):
                                st.warning("Debes ingresar el correo de todos los revisores asignados.")
                            elif nuevo_archivo and nueva_obs:
                                marca_tiempo = datetime.now(ZONA_COL).strftime("%H%M%S")
                                nuevo_nombre_nc = f"{doc_id_input}_v{marca_tiempo}_{nuevo_archivo.name}"
                                
                                with st.spinner("Subiendo corrección a Nextcloud..."):
                                    exito = subir_a_nextcloud(nuevo_archivo.getvalue(), nuevo_nombre_nc)
                                
                                if exito:
                                    c.execute("UPDATE firmas_flujo SET decision = decision || ' (Histórico)' WHERE doc_id = ? AND decision NOT LIKE '%(Histórico)'", (doc_id_input,))
                                    
                                    tokens_reinicio = []
                                    for correo_esp in correos_reinicio:
                                        tok = str(uuid.uuid4())
                                        c.execute("INSERT INTO firmas_flujo (doc_id, token, rol, correo_esperado, decision) VALUES (?, ?, 'REVISOR', ?, 'PENDIENTE')", (doc_id_input, tok, correo_esp.strip().lower()))
                                        tokens_reinicio.append(tok)
                                    
                                    obs_actualizada = f"{doc[4]} \n\n[REINICIO {datetime.now(ZONA_COL).strftime('%d/%m/%Y')}]: {nueva_obs}"
                                    
                                    c.execute("""UPDATE documentos 
                                                 SET estado = 'REVISION', nombre_archivo = ?, ruta_archivo = ?, 
                                                     observacion_inicial = ?, total_revisores = ? 
                                                 WHERE id = ?""", 
                                              (nuevo_archivo.name, nuevo_nombre_nc, obs_actualizada, nuevos_rev, doc_id_input))
                                    
                                    conn.commit()
                                    st.success("¡Flujo reiniciado exitosamente!")
                                    st.info("Envía estos nuevos enlaces a los revisores:")
                                    for tok in tokens_reinicio:
                                        st.code(f"{BASE_URL}/?token={tok}", language="text")
                                else:
                                    st.error("Error al subir el nuevo documento a Nextcloud.")
                            else:
                                st.warning("Debes adjuntar el documento corregido y escribir una nota.")
                else:
                    st.error("Documento no encontrado. Verifica el ID.")

    elif menu == "3. Documentos Finalizados":
        st.header("Repositorio de Documentos Completados")
        with sqlite3.connect(DB_FILE) as conn:
            c = conn.cursor()
            c.execute("""SELECT id, nombre_archivo, ruta_archivo, fecha_creacion, nombre_elaborador 
                         FROM documentos WHERE estado = 'COMPLETADO' ORDER BY fecha_creacion DESC""")
            docs = c.fetchall()
            
            if docs:
                for d in docs:
                    with st.expander(f"📄 {d[1]} (ID: {d[0]}) - Elaborado por: {d[4]}"):
                        st.write(f"**Fecha de inicio del flujo:** {d[3]}")
                        nombre_final_nc = d[2].replace(".pdf", "_FINAL.pdf")
                        st.write(f"**Nombre en Nextcloud:** `{nombre_final_nc}`")
                        
                        if st.button(f"Descargar de Nextcloud", key=f"btn_dl_nc_{d[0]}"):
                            archivo_bytes = descargar_de_nextcloud(nombre_final_nc)
                            if archivo_bytes:
                                st.download_button("📥 Confirmar Descarga", data=archivo_bytes, file_name=nombre_final_nc, mime="application/pdf", key=f"dl_{d[0]}")
                            else:
                                st.error("El archivo final ya no se encuentra en Nextcloud.")
            else:
                st.info("Aún no hay documentos con el flujo completado.")

    elif menu == "4. Respaldo de Base de Datos":
        st.header("Descargar Base de Datos (Auditoría)")
        st.write("Descarga el archivo SQLite actual para revisarlo con DB Browser.")
        
        if os.path.exists(DB_FILE):
            with open(DB_FILE, "rb") as db_file:
                st.download_button(
                    label="📥 Descargar flujo_documental.db",
                    data=db_file,
                    file_name=f"respaldo_bd_{datetime.now(ZONA_COL).strftime('%Y%m%d_%H%M')}.db",
                    mime="application/octet-stream",
                    key="btn_dl_db"
                )
        else:
            st.error("El archivo de base de datos no se encontró.")
